from flask import Flask, render_template, redirect, session, request, send_file, jsonify
import sqlite3
import os
import random
import smtplib
import threading
import csv
from io import BytesIO
from email.message import EmailMessage
from datetime import datetime, timedelta

from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from reportlab.pdfgen import canvas
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

import psycopg2
from psycopg2.extras import DictCursor

import cloudinary
import cloudinary.uploader

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env"))

app = Flask(__name__)

SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    SECRET_KEY = os.urandom(32)
    app.logger.warning("SECRET_KEY is not set. Generated a random secret key for this runtime.")
app.secret_key = SECRET_KEY

USE_HTTPS = os.getenv("USE_HTTPS", "0") == "1"
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=USE_HTTPS,
)

UPLOAD_FOLDER = "static/uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ADMIN_USERNAME = os.getenv("ADMIN_USERNAME")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")
if not ADMIN_USERNAME or not ADMIN_PASSWORD:
    app.logger.warning("ADMIN_USERNAME or ADMIN_PASSWORD not configured. Admin login is disabled until both are set.")

EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS")
EMAIL_APP_PASSWORD = os.getenv("EMAIL_APP_PASSWORD")

DEFAULT_COUPONS = {
    "SAVE10": 10,
}

cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET"),
    secure=True
)


def is_postgres():
    return bool(os.getenv("DATABASE_URL"))


def fix_sql(sql):
    """Convert SQLite ? placeholders to PostgreSQL %s placeholders."""
    if is_postgres():
        return sql.replace("?", "%s")
    return sql


def get_db():
    database_url = os.getenv("DATABASE_URL")

    if database_url:
        return psycopg2.connect(database_url, cursor_factory=DictCursor)

    conn = sqlite3.connect("grocery.db")
    conn.row_factory = sqlite3.Row
    return conn


def execute(cursor, sql, params=()):
    cursor.execute(fix_sql(sql), params)


def fetch_product_id(product):
    return product["id"] if hasattr(product, "keys") and "id" in product.keys() else product[0]


def product_col(product, key, index, default=""):
    try:
        if hasattr(product, "keys") and key in product.keys():
            return product[key]
    except Exception:
        pass

    try:
        return product[index]
    except Exception:
        return default


def product_stock(product):
    try:
        return int(product_col(product, "stock", 5, 0) or 0)
    except (TypeError, ValueError):
        return 0


def can_add_to_cart(product_id, current_quantity):
    conn = get_db()
    cursor = conn.cursor()
    execute(cursor, "SELECT * FROM products WHERE id=?", (product_id,))
    product = cursor.fetchone()
    conn.close()

    if not product:
        return False

    return product_stock(product) > current_quantity


def add_product_to_cart(product_id, quantity):
    cart = session.get("cart", {})
    if isinstance(cart, list):
        cart = {}

    quantity = safe_quantity(quantity)
    if quantity <= 0:
        quantity = 1

    cart_key = str(product_id)
    current_quantity = int(cart.get(cart_key, 0) or 0)
    added = 0

    for _ in range(quantity):
        if not can_add_to_cart(product_id, current_quantity + added):
            break
        added += 1

    if added > 0:
        cart[cart_key] = current_quantity + added
        session["cart"] = cart
        session.modified = True

    return added


def coupon_value(coupon):
    try:
        if hasattr(coupon, "keys") and "discount" in coupon.keys():
            return int(coupon["discount"] or 0)
    except Exception:
        pass

    try:
        return int(coupon[0] or 0)
    except Exception:
        return 0


def get_cart_items():
    cart_items = session.get("cart", {})
    if isinstance(cart_items, list) or not isinstance(cart_items, dict):
        session["cart"] = {}
        return {}
    return cart_items


def get_current_user_id():
    user_id = session.get("user_id")
    try:
        return int(user_id)
    except (TypeError, ValueError):
        return None


def get_session_wishlist():
    wishlist = session.get("wishlist", [])
    if not isinstance(wishlist, list):
        wishlist = []
    cleaned = []
    for item in wishlist:
        try:
            cleaned.append(str(int(item)))
        except (TypeError, ValueError):
            continue
    session["wishlist"] = cleaned
    session.modified = True
    return cleaned


def is_product_in_wishlist(product_id):
    return str(product_id) in get_session_wishlist()


def toggle_session_wishlist(product_id):
    wishlist = get_session_wishlist()
    product_key = str(product_id)
    if product_key in wishlist:
        wishlist.remove(product_key)
    else:
        wishlist.append(product_key)
    session["wishlist"] = wishlist
    session.modified = True


def safe_quantity(quantity):
    try:
        return max(0, int(quantity))
    except (TypeError, ValueError):
        return 0


def get_user_wishlist_ids():
    user_id = get_current_user_id()
    if not user_id:
        return {int(pid) for pid in get_session_wishlist() if str(pid).isdigit()}

    conn = get_db()
    cursor = conn.cursor()
    execute(cursor, "SELECT product_id FROM wishlist WHERE user_id=?", (user_id,))
    ids = {row[0] for row in cursor.fetchall()}
    conn.close()
    return ids


def toggle_user_wishlist(product_id):
    user_id = get_current_user_id()
    if not user_id:
        toggle_session_wishlist(product_id)
        return

    conn = get_db()
    cursor = conn.cursor()
    execute(cursor, "SELECT id FROM wishlist WHERE user_id=? AND product_id=?", (user_id, product_id))
    existing = cursor.fetchone()
    if existing:
        execute(cursor, "DELETE FROM wishlist WHERE id=?", (existing[0],))
    else:
        try:
            execute(cursor, "INSERT INTO wishlist(user_id, product_id) VALUES(?,?)", (user_id, product_id))
        except Exception:
            conn.rollback()

    conn.commit()
    conn.close()


def migrate_session_wishlist_to_user():
    user_id = get_current_user_id()
    if not user_id:
        return

    wishlist_ids = [int(pid) for pid in get_session_wishlist() if str(pid).isdigit()]
    if not wishlist_ids:
        session.pop("wishlist", None)
        return

    conn = get_db()
    cursor = conn.cursor()
    for product_id in wishlist_ids:
        try:
            execute(cursor, "INSERT INTO wishlist(user_id, product_id) VALUES(?,?)", (user_id, product_id))
        except Exception:
            conn.rollback()
            cursor = conn.cursor()
    conn.commit()
    conn.close()
    session.pop("wishlist", None)


def build_cart_summary():
    cart_items = get_cart_items()
    products = []
    total = 0
    order_products = []

    if not cart_items:
        return products, total, order_products

    conn = get_db()
    cursor = conn.cursor()

    try:
        for product_id, quantity in cart_items.items():
            quantity = safe_quantity(quantity)
            if quantity <= 0:
                continue

            execute(cursor, "SELECT * FROM products WHERE id=?", (product_id,))
            product = cursor.fetchone()

            if not product:
                continue

            price = int(product_col(product, "price", 3, 0) or 0)
            subtotal = price * quantity
            total += subtotal
            order_products.append((product, quantity))
            products.append({
                "name": product_col(product, "name", 1),
                "price": price,
                "quantity": quantity,
                "subtotal": subtotal,
            })
    finally:
        conn.close()

    return products, total, order_products


def order_totals(total):
    gst = int(total * 0.05)
    discount_percent = int(session.get("discount", 0) or 0)
    discount_amount = int(total * discount_percent / 100)
    grand_total = total + gst - discount_amount
    return gst, discount_percent, discount_amount, grand_total


def can_review_product(product_id):
    user = session.get("user")
    if not user:
        return False

    conn = get_db()
    cursor = conn.cursor()

    try:
        execute(cursor, "SELECT name FROM products WHERE id=?", (product_id,))
        product = cursor.fetchone()
        if not product:
            return False

        product_name = product_col(product, "name", 0)

        execute(
            cursor,
            """
            SELECT 1
            FROM orders
            JOIN order_items ON orders.id = order_items.order_id
            WHERE orders.status = ?
              AND (orders.customer_email = ? OR orders.customer_name = ?)
              AND (order_items.product_id = ? OR order_items.product_name = ?)
            LIMIT 1
            """,
            ("Delivered", user, user, product_id, product_name)
        )
        return cursor.fetchone() is not None
    finally:
        conn.close()


def empty_checkout_context():
    return {
        "products": [],
        "total": 0,
        "gst": 0,
        "discount_percent": 0,
        "discount_amount": 0,
        "grand_total": 0,
        "error": "",
        "form_data": {},
        "saved_addresses": [],
    }


def checkout_context(products, total, error="", form_data=None, saved_addresses=None):
    gst, discount_percent, discount_amount, grand_total = order_totals(total)
    return {
        "products": products,
        "total": total,
        "gst": gst,
        "discount_percent": discount_percent,
        "discount_amount": discount_amount,
        "grand_total": grand_total,
        "error": error,
        "form_data": form_data or {},
        "saved_addresses": saved_addresses or [],
    }


def current_address_user_key(email="", phone=""):
    return session.get("user") or session.get("last_checkout_user_key") or email or phone or ""


def ensure_saved_addresses_table(cursor):
    if is_postgres():
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS saved_addresses (
                id SERIAL PRIMARY KEY,
                user_key TEXT,
                name TEXT,
                phone TEXT,
                email TEXT,
                address TEXT,
                district TEXT,
                state TEXT,
                pincode TEXT
            )
        """)
    else:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS saved_addresses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_key TEXT,
                name TEXT,
                phone TEXT,
                email TEXT,
                address TEXT,
                district TEXT,
                state TEXT,
                pincode TEXT
            )
        """)

    try:
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_saved_addresses_user_key ON saved_addresses(user_key)")
    except Exception:
        pass


def fetch_saved_addresses(user_key):
    if not user_key:
        return []

    conn = None
    try:
        conn = get_db()
        cursor = conn.cursor()
        ensure_saved_addresses_table(cursor)
        conn.commit()
        execute(
            cursor,
            """
            SELECT id, name, phone, email, address, district, state, pincode
            FROM saved_addresses
            WHERE user_key=?
            ORDER BY id DESC
            """,
            (user_key,)
        )
        rows = cursor.fetchall()
    except Exception as error:
        app.logger.exception("Saved address fetch failed: %s", error)
        return []
    finally:
        if conn:
            conn.close()

    addresses = []
    for row in rows:
        addresses.append({
            "id": row[0],
            "name": row[1] or "",
            "phone": row[2] or "",
            "email": row[3] or "",
            "address": row[4] or "",
            "district": row[5] or "",
            "state": row[6] or "",
            "pincode": row[7] or "",
        })
    return addresses


def save_checkout_address(cursor, user_key, name, phone, email, address, district, state, pincode):
    if not user_key:
        return

    execute(
        cursor,
        """
        SELECT id FROM saved_addresses
        WHERE user_key=? AND address=? AND district=? AND state=? AND pincode=?
        LIMIT 1
        """,
        (user_key, address, district, state, pincode)
    )
    existing = cursor.fetchone()

    if existing:
        execute(
            cursor,
            """
            UPDATE saved_addresses
            SET name=?, phone=?, email=?
            WHERE id=?
            """,
            (name, phone, email, existing[0])
        )
        return

    execute(
        cursor,
        """
        INSERT INTO saved_addresses(user_key, name, phone, email, address, district, state, pincode)
        VALUES(?,?,?,?,?,?,?,?)
        """,
        (user_key, name, phone, email, address, district, state, pincode)
    )


def save_checkout_address_safe(user_key, name, phone, email, address, district, state, pincode):
    if not user_key:
        return

    conn = None
    try:
        conn = get_db()
        cursor = conn.cursor()
        ensure_saved_addresses_table(cursor)
        save_checkout_address(cursor, user_key, name, phone, email, address, district, state, pincode)
        conn.commit()
    except Exception as error:
        app.logger.exception("Saved address save failed: %s", error)
        if conn:
            conn.rollback()
    finally:
        if conn:
            conn.close()


def ensure_orders_created_at_column(cursor):
    if is_postgres():
        cursor.execute("SAVEPOINT ensure_orders_created_at_column")
        try:
            cursor.execute("ALTER TABLE orders ADD COLUMN created_at TEXT DEFAULT CURRENT_TIMESTAMP")
            cursor.execute("RELEASE SAVEPOINT ensure_orders_created_at_column")
        except Exception:
            cursor.execute("ROLLBACK TO SAVEPOINT ensure_orders_created_at_column")
            cursor.execute("RELEASE SAVEPOINT ensure_orders_created_at_column")
        return

    try:
        cursor.execute("ALTER TABLE orders ADD COLUMN created_at TEXT DEFAULT CURRENT_TIMESTAMP")
    except sqlite3.OperationalError:
        pass


def order_col(order, key, index, default=""):
    try:
        if hasattr(order, "keys") and key in order.keys():
            return order[key]
    except Exception:
        pass

    try:
        return order[index]
    except Exception:
        return default


def init_db():
    conn = get_db()
    cursor = conn.cursor()

    if is_postgres():
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id SERIAL PRIMARY KEY,
                name TEXT,
                category TEXT,
                price INTEGER,
                image TEXT,
                stock INTEGER,
                description TEXT,
                image2 TEXT,
                image3 TEXT,
                image4 TEXT,
                image5 TEXT,
                discount INTEGER DEFAULT 10
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS reviews (
                id SERIAL PRIMARY KEY,
                product_id INTEGER,
                customer_name TEXT,
                rating INTEGER,
                comment TEXT
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS wishlist (
                id SERIAL PRIMARY KEY,
                user_id INTEGER,
                product_id INTEGER,
                UNIQUE(user_id, product_id)
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                name TEXT,
                email TEXT UNIQUE,
                phone TEXT UNIQUE,
                password TEXT
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id SERIAL PRIMARY KEY,
                customer_name TEXT,
                phone TEXT,
                address TEXT,
                total INTEGER,
                payment_method TEXT,
                payment_status TEXT,
                gst_amount INTEGER,
                status TEXT,
                customer_email TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        ensure_orders_created_at_column(cursor)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS order_items (
                id SERIAL PRIMARY KEY,
                order_id INTEGER,
                product_id INTEGER,
                product_name TEXT,
                price INTEGER,
                quantity INTEGER
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS coupons (
                id SERIAL PRIMARY KEY,
                code TEXT UNIQUE,
                discount INTEGER
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS saved_addresses (
                id SERIAL PRIMARY KEY,
                user_key TEXT,
                name TEXT,
                phone TEXT,
                email TEXT,
                address TEXT,
                district TEXT,
                state TEXT,
                pincode TEXT
            )
        """)

        # Add columns safely for old deployed databases.
        extra_columns = {
            "products": [
                ("description", "TEXT"),
                ("image2", "TEXT"),
                ("image3", "TEXT"),
                ("image4", "TEXT"),
                ("image5", "TEXT"),
                ("discount", "INTEGER DEFAULT 10"),
            ],
            "orders": [
                ("status", "TEXT"),
                ("customer_email", "TEXT"),
                ("created_at", "TEXT DEFAULT CURRENT_TIMESTAMP"),
            ],
            "users": [
                ("phone", "TEXT"),
            ],
            "order_items": [
                ("product_id", "INTEGER"),
            ],
        }

        for table, columns in extra_columns.items():
            for name, col_type in columns:
                savepoint = f"alter_{table}_{name}"
                cursor.execute(f"SAVEPOINT {savepoint}")
                try:
                    cursor.execute(f"ALTER TABLE {table} ADD COLUMN {name} {col_type}")
                    cursor.execute(f"RELEASE SAVEPOINT {savepoint}")
                except Exception:
                    cursor.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
                    cursor.execute(f"RELEASE SAVEPOINT {savepoint}")
        cursor.execute("SAVEPOINT wishlist_user_id")
        try:
            cursor.execute("ALTER TABLE wishlist ADD COLUMN user_id INTEGER")
            cursor.execute("RELEASE SAVEPOINT wishlist_user_id")
        except Exception:
            cursor.execute("ROLLBACK TO SAVEPOINT wishlist_user_id")
            cursor.execute("RELEASE SAVEPOINT wishlist_user_id")
        cursor.execute("SAVEPOINT wishlist_user_product_index")
        try:
            cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_wishlist_user_product ON wishlist(user_id, product_id)")
            cursor.execute("RELEASE SAVEPOINT wishlist_user_product_index")
        except Exception:
            cursor.execute("ROLLBACK TO SAVEPOINT wishlist_user_product_index")
            cursor.execute("RELEASE SAVEPOINT wishlist_user_product_index")

        indexes = [
            "CREATE INDEX IF NOT EXISTS idx_products_category ON products(category)",
            "CREATE INDEX IF NOT EXISTS idx_reviews_product_id ON reviews(product_id)",
            "CREATE INDEX IF NOT EXISTS idx_orders_customer_email ON orders(customer_email)",
            "CREATE INDEX IF NOT EXISTS idx_order_items_order_id ON order_items(order_id)",
            "CREATE INDEX IF NOT EXISTS idx_order_items_product_id ON order_items(product_id)",
            "CREATE INDEX IF NOT EXISTS idx_saved_addresses_user_key ON saved_addresses(user_key)",
        ]

        for index_number, index_sql in enumerate(indexes):
            savepoint = f"index_savepoint_{index_number}"
            cursor.execute(f"SAVEPOINT {savepoint}")
            try:
                cursor.execute(index_sql)
                cursor.execute(f"RELEASE SAVEPOINT {savepoint}")
            except Exception:
                cursor.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
                cursor.execute(f"RELEASE SAVEPOINT {savepoint}")

    else:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                category TEXT,
                price INTEGER,
                image TEXT,
                stock INTEGER,
                description TEXT,
                image2 TEXT,
                image3 TEXT,
                image4 TEXT,
                image5 TEXT,
                discount INTEGER DEFAULT 10
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS reviews (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER,
                customer_name TEXT,
                rating INTEGER,
                comment TEXT
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS wishlist (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                product_id INTEGER,
                UNIQUE(user_id, product_id)
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                email TEXT UNIQUE,
                phone TEXT UNIQUE,
                password TEXT
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_name TEXT,
                phone TEXT,
                address TEXT,
                total INTEGER,
                payment_method TEXT,
                payment_status TEXT,
                gst_amount INTEGER,
                status TEXT,
                customer_email TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER,
                product_id INTEGER,
                product_name TEXT,
                price INTEGER,
                quantity INTEGER
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS coupons (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE,
                discount INTEGER
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS saved_addresses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_key TEXT,
                name TEXT,
                phone TEXT,
                email TEXT,
                address TEXT,
                district TEXT,
                state TEXT,
                pincode TEXT
            )
        """)

        extra_columns = {
            "products": [
                ("description", "TEXT"),
                ("image2", "TEXT"),
                ("image3", "TEXT"),
                ("image4", "TEXT"),
                ("image5", "TEXT"),
                ("discount", "INTEGER DEFAULT 10"),
            ],
            "orders": [
                ("status", "TEXT"),
                ("customer_email", "TEXT"),
                ("created_at", "TEXT DEFAULT CURRENT_TIMESTAMP"),
            ],
            "users": [
                ("phone", "TEXT"),
            ],
            "order_items": [
                ("product_id", "INTEGER"),
            ],
        }

        for table, columns in extra_columns.items():
            for name, col_type in columns:
                try:
                    cursor.execute(f"ALTER TABLE {table} ADD COLUMN {name} {col_type}")
                except sqlite3.OperationalError:
                    pass
        try:
            cursor.execute("ALTER TABLE wishlist ADD COLUMN user_id INTEGER")
        except sqlite3.OperationalError:
            pass
        try:
            cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_wishlist_user_product ON wishlist(user_id, product_id)")
        except sqlite3.OperationalError:
            pass

        indexes = [
            "CREATE INDEX IF NOT EXISTS idx_products_category ON products(category)",
            "CREATE INDEX IF NOT EXISTS idx_reviews_product_id ON reviews(product_id)",
            "CREATE INDEX IF NOT EXISTS idx_orders_customer_email ON orders(customer_email)",
            "CREATE INDEX IF NOT EXISTS idx_order_items_order_id ON order_items(order_id)",
            "CREATE INDEX IF NOT EXISTS idx_order_items_product_id ON order_items(product_id)",
            "CREATE INDEX IF NOT EXISTS idx_saved_addresses_user_key ON saved_addresses(user_key)",
        ]

        for index_sql in indexes:
            cursor.execute(index_sql)

    conn.commit()
    conn.close()


def ensure_product_columns():
    init_db()


def insert_product_record(cursor, name, category, price, stock, description="", images=None):
    images = (images or [])[:5]
    image1 = images[0] if len(images) > 0 else ""
    image2 = images[1] if len(images) > 1 else ""
    image3 = images[2] if len(images) > 2 else ""
    image4 = images[3] if len(images) > 3 else ""
    image5 = images[4] if len(images) > 4 else ""

    execute(
        cursor,
        """
        INSERT INTO products
        (name, category, price, image, stock, description, image2, image3, image4, image5)
        VALUES (?,?,?,?,?,?,?,?,?,?)
        """,
        (
            name,
            category,
            price,
            image1,
            stock,
            description,
            image2,
            image3,
            image4,
            image5
        )
    )


init_db()


def send_email(to_email, subject, body):
    if not EMAIL_ADDRESS or not EMAIL_APP_PASSWORD:
        print("Email settings missing")
        return False, "Email settings missing. Render me EMAIL_ADDRESS aur EMAIL_APP_PASSWORD set kare."

    try:
        msg = EmailMessage()
        msg["From"] = EMAIL_ADDRESS
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.set_content(body)

        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=8) as smtp:
            smtp.login(EMAIL_ADDRESS, EMAIL_APP_PASSWORD)
            smtp.send_message(msg)
        return True, "OTP email sent."
    except Exception as e:
        print("Email send error:", e)
        return False, f"Email send failed: {e}"


def send_email_async(to_email, subject, body):
    def runner():
        send_email(to_email, subject, body)

    thread = threading.Thread(target=runner, daemon=True)
    thread.start()


def send_otp_email(to_email, subject, body):
    sent, message = send_email(to_email, subject, body)
    return sent, message


def is_admin():
    return session.get("admin") is True


def make_otp():
    return str(random.randint(100000, 999999))


def send_phone_otp(phone, otp):
    # SMS gateway nahi laga hai, isliye phone OTP demo mode me verify page par dikhega.
    print(f"Phone OTP for {phone}: {otp}")


def normalize_login_id(value):
    login_id = (value or "").strip().lower()
    if "@" in login_id:
        if "." not in login_id:
            return "", "", "Valid email enter kare."
        return login_id, "email", ""

    if not login_id.isdigit() or len(login_id) != 10:
        return "", "", "Phone number 10 digit ka hona chahiye ya valid email enter kare."

    return login_id, "phone", ""


def register_pending_user():
    pending = session.get("pending_register")
    if not pending:
        return None

    ensure_user_phone_column()
    hashed_password = generate_password_hash(pending["password"])

    conn = get_db()
    cursor = conn.cursor()

    try:
        if is_postgres():
            execute(
                cursor,
                "INSERT INTO users(name, email, phone, password) VALUES(?,?,?,?) RETURNING id",
                (pending["name"], pending.get("email"), pending.get("phone"), hashed_password)
            )
            user_id = cursor.fetchone()[0]
        else:
            execute(
                cursor,
                "INSERT INTO users(name, email, phone, password) VALUES(?,?,?,?)",
                (pending["name"], pending.get("email"), pending.get("phone"), hashed_password)
            )
            user_id = cursor.lastrowid
        conn.commit()
    except Exception:
        conn.rollback()
        conn.close()
        return None

    conn.close()
    pending["id"] = user_id
    return pending


def ensure_user_phone_column():
    conn = get_db()
    cursor = conn.cursor()

    try:
        execute(cursor, "ALTER TABLE users ADD COLUMN phone TEXT")
        conn.commit()
    except Exception:
        conn.rollback()
    finally:
        conn.close()


@app.route("/")
def home():
    search = request.args.get("search", "")
    category = request.args.get("category", "")
    stock_filter = request.args.get("stock", "")
    sort = request.args.get("sort", "")

    conn = get_db()
    cursor = conn.cursor()

    query = "SELECT * FROM products WHERE 1=1"
    params = []

    if search:
        query += " AND name LIKE ?"
        params.append("%" + search + "%")

    if category:
        query += " AND category=?"
        params.append(category)

    if stock_filter == "in":
        query += " AND stock > 0"
    elif stock_filter == "out":
        query += " AND stock <= 0"

    sort_options = {
        "price_low": "price ASC",
        "price_high": "price DESC",
        "stock_low": "stock ASC",
        "stock_high": "stock DESC",
        "newest": "id DESC",
    }
    query += " ORDER BY " + sort_options.get(sort, "id DESC")

    execute(cursor, query, params)
    products = cursor.fetchall()

    ratings = {}
    product_ids = [fetch_product_id(product) for product in products]

    if product_ids:
        placeholders = ",".join(["?"] * len(product_ids))
        execute(
            cursor,
            f"""
            SELECT product_id, AVG(rating), COUNT(*)
            FROM reviews
            WHERE product_id IN ({placeholders})
            GROUP BY product_id
            """,
            product_ids
        )

        for product_id, avg_rating, total_reviews in cursor.fetchall():
            ratings[product_id] = {
                "avg": round(avg_rating, 1) if avg_rating else 0,
                "count": total_reviews
            }

    cursor.execute("SELECT DISTINCT category FROM products")
    categories = cursor.fetchall()

    wishlist_ids = set()
    for raw_id in get_session_wishlist():
        try:
            wishlist_ids.add(int(raw_id))
        except ValueError:
            continue

    conn.close()

    return render_template(
        "index.html",
        products=products,
        ratings=ratings,
        search=search,
        category=category,
        stock_filter=stock_filter,
        sort=sort,
        categories=categories,
        wishlist_ids=wishlist_ids
    )


@app.route("/send_otp", methods=["GET", "POST"])
def send_otp():
    ensure_user_phone_column()

    if request.method == "POST":
        login_id, login_type, login_error = normalize_login_id(request.form.get("login_id", ""))
        otp = make_otp()

        if login_error:
            return render_template("send_otp.html", error=login_error, form_data=request.form)

        conn = get_db()
        cursor = conn.cursor()
        execute(cursor, "SELECT * FROM users WHERE LOWER(email)=? OR phone=?", (login_id, login_id))
        user = cursor.fetchone()
        conn.close()

        if not user:
            return render_template("send_otp.html", error="Is email ya phone se account nahi mila.", form_data=request.form)

        user_email = user["email"] if hasattr(user, "keys") and "email" in user.keys() else user[2]
        user_phone = user["phone"] if hasattr(user, "keys") and "phone" in user.keys() else ""
        user_id = user["id"] if hasattr(user, "keys") and "id" in user.keys() else user[0]

        session["otp_user"] = user_email or user_phone
        session["otp_user_id"] = user_id
        session["otp_login_id"] = login_id
        session["otp"] = otp
        session["otp_flow"] = "login"

        session.pop("otp_notice", None)
        if login_type == "email":
            sent, message = send_otp_email(
                user_email,
                "Your Damodar Grocery Store OTP",
                f"Your login OTP is {otp}"
            )
            session["phone_demo_otp"] = ""
            session["email_demo_otp"] = "" if sent else otp
            if not sent:
                session["otp_notice"] = message
        else:
            send_phone_otp(user_phone, otp)
            session["phone_demo_otp"] = otp
            session["email_demo_otp"] = ""

        return redirect("/verify_otp")

    return render_template("send_otp.html", form_data={})


@app.route("/verify_otp", methods=["GET", "POST"])
def verify_otp():
    if request.method == "POST":
        user_otp = request.form.get("otp", "").strip()

        if session.get("otp_flow") == "register":
            if user_otp == session.get("otp"):
                pending = register_pending_user()
                if not pending:
                    session.pop("otp", None)
                    return render_template(
                        "verify_otp.html",
                        error="Account create nahi ho paya. Email ya phone already registered ho sakta hai."
                    )

                session["user"] = pending.get("email") or pending.get("phone")
                session["user_id"] = pending.get("id") if pending else None
                session.pop("pending_register", None)

                session.pop("otp", None)
                session.pop("otp_flow", None)
                session.pop("phone_demo_otp", None)
                session.pop("email_demo_otp", None)
                session.pop("otp_notice", None)
                session.pop("admin", None)
                migrate_session_wishlist_to_user()
                return redirect("/")

            return render_template(
                "verify_otp.html",
                error="OTP galat hai.",
                email_demo_otp=session.get("email_demo_otp", ""),
                phone_demo_otp=session.get("phone_demo_otp", ""),
                otp_notice=session.get("otp_notice", ""),
                otp_flow="register"
            )

        if user_otp == session.get("otp"):
            session["user"] = session.get("otp_user")
            session["user_id"] = session.get("otp_user_id")
            session.pop("otp", None)
            session.pop("otp_flow", None)
            session.pop("otp_user", None)
            session.pop("otp_user_id", None)
            session.pop("otp_login_id", None)
            session.pop("phone_demo_otp", None)
            session.pop("email_demo_otp", None)
            session.pop("otp_notice", None)
            session.pop("admin", None)
            migrate_session_wishlist_to_user()
            return redirect("/")

        return render_template(
            "verify_otp.html",
            error="Wrong OTP. Dobara check kare.",
            email_demo_otp=session.get("email_demo_otp", ""),
            phone_demo_otp=session.get("phone_demo_otp", ""),
            otp_notice=session.get("otp_notice", ""),
            otp_flow=session.get("otp_flow", "login")
        )

    return render_template(
        "verify_otp.html",
        email_demo_otp=session.get("email_demo_otp", ""),
        phone_demo_otp=session.get("phone_demo_otp", ""),
        otp_notice=session.get("otp_notice", ""),
        otp_flow=session.get("otp_flow", "login")
    )


@app.route("/coupons", methods=["GET", "POST"])
def coupons():
    if not is_admin():
        return redirect("/admin_login")

    message = ""
    error = ""
    conn = get_db()
    cursor = conn.cursor()

    if request.method == "POST":
        code = request.form.get("code", "").strip().upper()
        discount = request.form.get("discount", "").strip()

        try:
            discount_value = int(discount)
        except ValueError:
            discount_value = 0

        if not code:
            error = "Coupon code required hai."
        elif discount_value < 1 or discount_value > 100:
            error = "Discount 1 se 100 percent ke beech hona chahiye."
        else:
            try:
                execute(cursor, "INSERT INTO coupons(code, discount) VALUES(?,?)", (code, discount_value))
                conn.commit()
                message = f"Coupon {code} add ho gaya."
            except Exception:
                conn.rollback()
                cursor = conn.cursor()
                execute(cursor, "UPDATE coupons SET discount=? WHERE code=?", (discount_value, code))
                conn.commit()
                message = f"Coupon {code} update ho gaya."

    execute(cursor, "SELECT * FROM coupons ORDER BY id DESC")
    all_coupons = cursor.fetchall()
    conn.close()
    return render_template("coupons.html", coupons=all_coupons, success=message, error=error)


@app.route("/edit_coupon/<int:coupon_id>", methods=["POST"])
def edit_coupon(coupon_id):
    if not is_admin():
        return redirect("/admin_login")

    code = request.form.get("code", "").strip().upper()
    discount = request.form.get("discount", "").strip()

    try:
        discount_value = int(discount)
    except ValueError:
        discount_value = 0

    if not code or discount_value < 1 or discount_value > 100:
        return redirect("/coupons")

    conn = get_db()
    cursor = conn.cursor()
    try:
        execute(cursor, "UPDATE coupons SET code=?, discount=? WHERE id=?", (code, discount_value, coupon_id))
        conn.commit()
    except Exception:
        conn.rollback()
    conn.close()
    return redirect("/coupons")


@app.route("/delete_coupon/<int:coupon_id>", methods=["POST"])
def delete_coupon(coupon_id):
    if not is_admin():
        return redirect("/admin_login")

    conn = get_db()
    cursor = conn.cursor()
    execute(cursor, "DELETE FROM coupons WHERE id=?", (coupon_id,))
    conn.commit()
    conn.close()
    return redirect("/coupons")

@app.route("/apply_coupon", methods=["POST"])
def apply_coupon():
    code = request.form["coupon"].upper().strip()

    if not code:
        session["coupon_code"] = ""
        session["discount"] = 0
        session["coupon_message"] = "Please enter a coupon code."
        session["coupon_status"] = "error"
        return redirect("/cart")

    conn = get_db()
    cursor = conn.cursor()

    execute(cursor, "SELECT discount FROM coupons WHERE code=?", (code,))
    coupon = cursor.fetchone()

    conn.close()

    if coupon:
        session["coupon_code"] = code
        session["discount"] = coupon_value(coupon)
        session["coupon_message"] = f"Coupon {code} applied successfully."
        session["coupon_status"] = "success"
    elif code in DEFAULT_COUPONS:
        session["coupon_code"] = code
        session["discount"] = DEFAULT_COUPONS[code]
        session["coupon_message"] = f"Coupon {code} applied successfully."
        session["coupon_status"] = "success"
    else:
        session["coupon_code"] = ""
        session["discount"] = 0
        session["coupon_message"] = "Invalid coupon code."
        session["coupon_status"] = "error"

    return redirect("/cart")


@app.route("/remove_coupon")
def remove_coupon():
    session["coupon_code"] = ""
    session["discount"] = 0
    session["coupon_message"] = "Coupon removed."
    session["coupon_status"] = "success"
    return redirect("/cart")


@app.route("/review/<int:product_id>", methods=["GET", "POST"])
def review(product_id):
    if not can_review_product(product_id):
        if "user" not in session:
            return redirect("/customer_login")
        return "Review only delivered order ke baad diya ja sakta hai."

    if request.method == "POST":
        customer_name = request.form["customer_name"]
        rating = request.form["rating"]
        comment = request.form["comment"]

        conn = get_db()
        cursor = conn.cursor()

        execute(
            cursor,
            "INSERT INTO reviews(product_id, customer_name, rating, comment) VALUES(?,?,?,?)",
            (product_id, customer_name, rating, comment)
        )

        conn.commit()
        conn.close()

        return redirect("/reviews/" + str(product_id))

    return render_template("review.html", product_id=product_id)


@app.route("/reviews/<int:product_id>")
def reviews(product_id):
    conn = get_db()
    cursor = conn.cursor()

    execute(cursor, "SELECT * FROM reviews WHERE product_id=?", (product_id,))
    reviews_data = cursor.fetchall()

    conn.close()

    return render_template("reviews.html", reviews=reviews_data)


@app.route("/profile")
def profile():
    if "user" not in session:
        return redirect("/customer_login")

    conn = get_db()
    cursor = conn.cursor()
    user_email_or_name = session["user"]
    execute(
        cursor,
        "SELECT * FROM orders WHERE customer_email=? OR customer_name=?",
        (user_email_or_name, user_email_or_name)
    )
    orders = cursor.fetchall()
    conn.close()

    total_orders = len(orders)
    delivered_orders = sum(1 for order in orders if order_col(order, "status", 8) == "Delivered")
    pending_orders = total_orders - delivered_orders

    return render_template(
        "profile.html",
        name=session["user"],
        total_orders=total_orders,
        delivered_orders=delivered_orders,
        pending_orders=pending_orders
    )


@app.route("/my_orders")
def my_orders():
    if "user" not in session:
        return redirect("/customer_login")

    conn = get_db()
    cursor = conn.cursor()

    user_identifier = str(session.get("user", "")).strip()
    identifiers = {"emails": set(), "phones": set(), "names": set()}

    if user_identifier:
        if "@" in user_identifier:
            identifiers["emails"].add(user_identifier.lower())
        elif user_identifier.isdigit():
            identifiers["phones"].add(user_identifier)
        else:
            identifiers["names"].add(user_identifier.lower())

    user_id = session.get("user_id")
    if user_id:
        execute(cursor, "SELECT name, email, phone FROM users WHERE id=?", (user_id,))
        user_row = cursor.fetchone()
        if user_row:
            user_name = order_col(user_row, "name", 0, "")
            user_email = order_col(user_row, "email", 1, "")
            user_phone = order_col(user_row, "phone", 2, "")
            if user_name:
                identifiers["names"].add(str(user_name).lower())
            if user_email:
                identifiers["emails"].add(str(user_email).lower())
            if user_phone:
                identifiers["phones"].add(str(user_phone))

    clauses = []
    params = []
    if identifiers["emails"]:
        placeholders = ",".join("?" for _ in identifiers["emails"])
        clauses.append(f"LOWER(customer_email) IN ({placeholders})")
        params.extend(sorted(identifiers["emails"]))
    if identifiers["phones"]:
        placeholders = ",".join("?" for _ in identifiers["phones"])
        clauses.append(f"phone IN ({placeholders})")
        params.extend(sorted(identifiers["phones"]))
    if identifiers["names"]:
        placeholders = ",".join("?" for _ in identifiers["names"])
        clauses.append(f"LOWER(customer_name) IN ({placeholders})")
        params.extend(sorted(identifiers["names"]))

    last_order_id = session.get("last_order_id")
    if last_order_id:
        clauses.append("id=?")
        params.append(last_order_id)

    if clauses:
        execute(cursor, f"SELECT * FROM orders WHERE {' OR '.join(clauses)} ORDER BY id DESC", tuple(params))
        orders = cursor.fetchall()
    else:
        orders = []

    order_items_map = {}
    order_meta_map = {}
    for order in orders:
        order_id = order_col(order, "id", 0)
        placed_raw = order_col(order, "created_at", 10, "")
        placed_meta = format_order_datetime(placed_raw)

        execute(cursor, "SELECT * FROM order_items WHERE order_id=?", (order_id,))
        items = []
        item_count = 0
        for item in cursor.fetchall():
            product_id = order_col(item, "product_id", 5, "")
            quantity = int(order_col(item, "quantity", 4, 0) or 0)
            image = ""
            if product_id:
                execute(cursor, "SELECT image FROM products WHERE id=?", (product_id,))
                product_row = cursor.fetchone()
                if product_row:
                    image = product_row[0] or ""

            item_count += quantity
            items.append({
                "product_id": product_id,
                "product_name": order_col(item, "product_name", 2, ""),
                "price": order_col(item, "price", 3, 0),
                "quantity": quantity,
                "image": image,
            })
        order_items_map[order_id] = items
        order_meta_map[order_id] = {
            "item_count": item_count,
            "placed_at": placed_meta["placed_at"],
            "placed_date": placed_meta["placed_date"],
            "placed_time": placed_meta["placed_time"],
            "expected_delivery": "2-4 days",
        }

    conn.close()

    return render_template(
        "my_orders.html",
        orders=orders,
        order_items_map=order_items_map,
        order_meta_map=order_meta_map
    )


@app.route("/admin_login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        if not ADMIN_USERNAME or not ADMIN_PASSWORD:
            return render_template("login.html", error="Admin login not configured.")

        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session.clear()
            session["admin"] = True
            return redirect("/orders")
        return render_template("login.html", error="Admin username ya password galat hai.")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


@app.route("/register", methods=["GET", "POST"])
def register():
    ensure_user_phone_column()

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        login_id, login_type, login_error = normalize_login_id(request.form.get("login_id", ""))
        password = request.form.get("password", "")

        if not name:
            return render_template("register.html", error="Name required hai.", form_data=request.form)

        if login_error:
            return render_template("register.html", error=login_error, form_data=request.form)

        if len(password) < 6:
            return render_template("register.html", error="Password kam se kam 6 characters ka hona chahiye.", form_data=request.form)

        conn = get_db()
        cursor = conn.cursor()
        execute(cursor, "SELECT 1 FROM users WHERE LOWER(email)=? OR phone=?", (login_id, login_id))
        existing_user = cursor.fetchone()
        conn.close()

        if existing_user:
            return render_template("register.html", error="Ye email ya phone already registered hai.", form_data=request.form)

        otp = make_otp()

        session["pending_register"] = {
            "name": name,
            "email": login_id if login_type == "email" else None,
            "phone": login_id if login_type == "phone" else None,
            "password": password,
        }
        session["otp"] = otp
        session["otp_flow"] = "register"

        session.pop("otp_notice", None)
        if login_type == "email":
            session["phone_demo_otp"] = ""
            sent, message = send_otp_email(
                login_id,
                "Verify your Damodar Grocery account",
                f"Your account verification OTP is {otp}"
            )
            session["email_demo_otp"] = "" if sent else otp
            if not sent:
                session["otp_notice"] = message
        else:
            session["email_demo_otp"] = ""
            session["phone_demo_otp"] = otp
            send_phone_otp(login_id, otp)

        return redirect("/verify_otp")

    return render_template("register.html", form_data={})


@app.route("/customer_login", methods=["GET", "POST"])
def customer_login():
    ensure_user_phone_column()

    if request.method == "POST":
        login_id, _, login_error = normalize_login_id(request.form.get("login_id", ""))
        password = request.form.get("password", "")

        if login_error:
            return render_template("customer_login.html", error=login_error, form_data=request.form)

        conn = get_db()
        cursor = conn.cursor()
        execute(cursor, "SELECT * FROM users WHERE LOWER(email)=? OR phone=?", (login_id, login_id))
        user = cursor.fetchone()
        conn.close()

        if user:
            stored_password = user["password"] if hasattr(user, "keys") and "password" in user.keys() else user[3]
            user_email = user["email"] if hasattr(user, "keys") and "email" in user.keys() else user[2]
            user_phone = user["phone"] if hasattr(user, "keys") and "phone" in user.keys() else ""
            user_id = user["id"] if hasattr(user, "keys") and "id" in user.keys() else user[0]

            # Supports old plain-text passwords and upgrades them after login.
            if check_password_hash(stored_password, password) or stored_password == password:
                session.pop("admin", None)
                session["user"] = user_email or user_phone
                session["user_id"] = user_id

                if stored_password == password:
                    conn = get_db()
                    cursor = conn.cursor()
                    execute(
                        cursor,
                        "UPDATE users SET password=? WHERE LOWER(email)=? OR phone=?",
                        (generate_password_hash(password), (user_email or "").lower(), user_phone)
                    )
                    conn.commit()
                    conn.close()

                return redirect("/")

        return render_template("customer_login.html", error="Email/phone ya password galat hai.", form_data=request.form)

    return render_template("customer_login.html", form_data={})


@app.route("/add_to_cart/<int:product_id>")
def add_to_cart(product_id):
    quantity = request.args.get("quantity", request.form.get("quantity", 1))
    add_product_to_cart(product_id, quantity)

    return redirect("/cart")


@app.route("/buy_now/<int:product_id>")
def buy_now(product_id):
    quantity = request.args.get("quantity", 1)
    add_product_to_cart(product_id, quantity)
    return redirect("/checkout")


@app.route("/increase/<int:product_id>")
def increase(product_id):
    cart = session.get("cart", {})
    if isinstance(cart, list):
        cart = {}

    cart_key = str(product_id)
    current_quantity = int(cart.get(cart_key, 0))

    if not can_add_to_cart(product_id, current_quantity):
        return redirect("/cart")

    cart[cart_key] = current_quantity + 1

    session["cart"] = cart
    session.modified = True

    return redirect("/cart")


@app.route("/decrease/<int:product_id>")
def decrease(product_id):
    cart = session.get("cart", {})
    if isinstance(cart, list):
        cart = {}

    product_id = str(product_id)

    if product_id in cart:
        cart[product_id] -= 1
        if cart[product_id] <= 0:
            cart.pop(product_id)

    session["cart"] = cart
    session.modified = True

    return redirect("/cart")


@app.route("/remove_from_cart/<int:product_id>")
def remove_from_cart(product_id):
    cart = session.get("cart", {})
    if isinstance(cart, list):
        cart = {}

    cart.pop(str(product_id), None)
    session["cart"] = cart
    session.modified = True

    if not cart:
        session["coupon_code"] = ""
        session["discount"] = 0
        session["coupon_message"] = ""
        session["coupon_status"] = ""

    return redirect("/cart")


@app.route("/cart")
def cart():
    cart_items = session.get("cart", {})
    if isinstance(cart_items, list):
        cart_items = {}
        session["cart"] = {}

    conn = get_db()
    cursor = conn.cursor()

    products = []
    total = 0

    for product_id, quantity in cart_items.items():
        execute(cursor, "SELECT * FROM products WHERE id=?", (product_id,))
        product = cursor.fetchone()

        if product:
            price = int(product_col(product, "price", 3, 0))
            subtotal = price * quantity
            total += subtotal
            products.append((product, quantity, subtotal))

    gst = int(total * 0.05)
    discount_percent = int(session.get("discount", 0) or 0)
    discount_amount = int(total * discount_percent / 100)
    grand_total = total + gst - discount_amount

    conn.close()

    item_count = sum(quantity for _, quantity, _ in products)

    return render_template(
        "cart.html",
        products=products,
        total=total,
        gst=gst,
        discount_percent=discount_percent,
        discount_amount=discount_amount,
        grand_total=grand_total,
        item_count=item_count,
        coupon_code=session.get("coupon_code", ""),
        coupon_message=session.get("coupon_message", ""),
        coupon_status=session.get("coupon_status", "")
    )


@app.route("/product/<int:product_id>")
def product_details(product_id):
    conn = get_db()
    cursor = conn.cursor()

    execute(cursor, "SELECT * FROM products WHERE id=?", (product_id,))
    product = cursor.fetchone()

    if not product:
        conn.close()
        return "Product not found"

    execute(
        cursor,
        "SELECT AVG(rating), COUNT(*) FROM reviews WHERE product_id=?",
        (product_id,)
    )
    data = cursor.fetchone()

    avg_rating = round(data[0], 1) if data and data[0] else 0
    total_reviews = data[1] if data else 0

    execute(cursor, "SELECT * FROM reviews WHERE product_id=? ORDER BY id DESC", (product_id,))
    reviews_data = cursor.fetchall()

    wishlist_ids = get_user_wishlist_ids()

    execute(
        cursor,
        "SELECT * FROM products WHERE category=? AND id!=? LIMIT 8",
        (product_col(product, "category", 2), product_id)
    )
    related_products = list(cursor.fetchall())

    if len(related_products) < 6:
        related_ids = {fetch_product_id(item) for item in related_products}
        related_ids.add(product_id)
        execute(cursor, "SELECT * FROM products WHERE id!=? LIMIT 12", (product_id,))
        for item in cursor.fetchall():
            item_id = fetch_product_id(item)
            if item_id not in related_ids:
                related_products.append(item)
                related_ids.add(item_id)
            if len(related_products) >= 8:
                break

    conn.close()

    return render_template(
        "product_details.html",
        product=product,
        avg_rating=avg_rating,
        total_reviews=total_reviews,
        reviews=reviews_data,
        wishlist_ids=wishlist_ids,
        related_products=related_products,
        can_review=can_review_product(product_id)
    )


@app.route("/wishlist/<int:product_id>")
def wishlist(product_id):
    toggle_user_wishlist(product_id)
    return redirect(request.referrer or "/")


@app.route("/wishlist")
def wishlist_page():
    user_id = get_current_user_id()
    products = []

    if user_id:
        conn = get_db()
        cursor = conn.cursor()
        execute(cursor, "SELECT products.* FROM wishlist JOIN products ON wishlist.product_id = products.id WHERE wishlist.user_id=?", (user_id,))
        products = cursor.fetchall()
        conn.close()
    else:
        wishlist_ids = get_session_wishlist()
        params = [int(pid) for pid in wishlist_ids if pid.isdigit()]
        if params:
            conn = get_db()
            cursor = conn.cursor()
            placeholders = ",".join(["?"] * len(params))
            execute(cursor, f"SELECT * FROM products WHERE id IN ({placeholders})", params)
            products = cursor.fetchall()
            conn.close()

    return render_template("wishlist.html", products=products)


@app.route("/checkout", methods=["GET", "POST"])
def checkout():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        phone = request.form.get("phone", "").strip()
        address = request.form.get("address", "").strip()
        district = request.form.get("district", "").strip()
        state = request.form.get("state", "").strip()
        pincode = request.form.get("pincode", "").strip()
        payment_method = request.form.get("payment_method", "COD")
        email = request.form.get("email", "").strip()
        full_address = f"{address}, District: {district}, State: {state}, Pincode: {pincode}"
        user_key = current_address_user_key(email, phone)
        saved_addresses = fetch_saved_addresses(user_key)

        products, total, order_products = build_cart_summary()

        if not order_products:
            return redirect("/cart")

        errors = []
        if not name:
            errors.append("Name required hai.")
        if not phone.isdigit() or len(phone) != 10:
            errors.append("Mobile number 10 digit ka hona chahiye.")
        if not email or "@" not in email or "." not in email:
            errors.append("Valid email enter kare.")
        if not address:
            errors.append("Full address required hai.")
        if not district:
            errors.append("District required hai.")
        if not state:
            errors.append("State required hai.")
        if not pincode.isdigit() or len(pincode) != 6:
            errors.append("Pin code 6 digit ka hona chahiye.")

        if errors:
            return render_template(
                "checkout.html",
                **checkout_context(products, total, " ".join(errors), request.form, saved_addresses)
            )

        for product, quantity in order_products:
            if product_stock(product) < quantity:
                return render_template(
                    "checkout.html",
                    **checkout_context(
                        products,
                        total,
                        f"{product_col(product, 'name', 1)} ke liye enough stock nahi hai.",
                        request.form,
                        saved_addresses
                    )
                )

        gst, discount_percent, discount_amount, grand_total = order_totals(total)
        order_created_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")

        conn = get_db()
        cursor = conn.cursor()
        ensure_orders_created_at_column(cursor)

        should_save_address = request.form.get("save_address") == "1"

        if is_postgres():
            execute(
                cursor,
                """
                INSERT INTO orders(customer_name, phone, address, total, payment_method, payment_status, gst_amount, status, customer_email, created_at)
                VALUES(?,?,?,?,?,?,?,?,?,?)
                RETURNING id
                """,
                (
                    name,
                    phone,
                    full_address,
                    grand_total,
                    payment_method,
                    "Paid" if payment_method == "Demo Payment" else "Pending",
                    gst,
                    "Pending",
                    email,
                    order_created_at
                )
            )
            order_id = cursor.fetchone()[0]
        else:
            execute(
                cursor,
                """
                INSERT INTO orders(customer_name, phone, address, total, payment_method, payment_status, gst_amount, status, customer_email, created_at)
                VALUES(?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    name,
                    phone,
                    full_address,
                    grand_total,
                    payment_method,
                    "Paid" if payment_method == "Demo Payment" else "Pending",
                    gst,
                    "Pending",
                    email,
                    order_created_at
                )
            )
            order_id = cursor.lastrowid

        for product, quantity in order_products:
            product_id = fetch_product_id(product)

            execute(
                cursor,
                "UPDATE products SET stock = stock - ? WHERE id=? AND stock >= ?",
                (quantity, product_id, quantity)
            )

            execute(
                cursor,
                "INSERT INTO order_items(order_id, product_id, product_name, price, quantity) VALUES(?,?,?,?,?)",
                (
                    order_id,
                    product_id,
                    product_col(product, "name", 1),
                    product_col(product, "price", 3),
                    quantity
                )
            )

        conn.commit()
        conn.close()

        if should_save_address:
            save_checkout_address_safe(user_key, name, phone, email, address, district, state, pincode)
            session["last_checkout_user_key"] = user_key

        send_email_async(
            email,
            "Order Placed Successfully",
            f"Your order #{order_id} has been placed. Total amount: Rs.{grand_total}"
        )

        session["cart"] = {}
        session["discount"] = 0
        session["coupon_code"] = ""
        session["coupon_message"] = ""
        session["coupon_status"] = ""
        session["last_order_id"] = order_id

        return render_template("success.html", order_id=order_id)

    try:
        products, total, order_products = build_cart_summary()

        if not order_products:
            return redirect("/cart")

        user_key = current_address_user_key()
        return render_template(
            "checkout.html",
            **checkout_context(products, total, saved_addresses=fetch_saved_addresses(user_key))
        )
    except Exception as error:
        app.logger.exception("Checkout page error: %s", error)
        return render_template("checkout.html", **empty_checkout_context())


@app.route("/invoice/<int:order_id>")
def invoice(order_id):
    conn = get_db()
    cursor = conn.cursor()

    execute(cursor, "SELECT * FROM orders WHERE id=?", (order_id,))
    order = cursor.fetchone()

    execute(cursor, "SELECT * FROM order_items WHERE order_id=?", (order_id,))
    items = cursor.fetchall()

    conn.close()

    if not order:
        return "Order not found"

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer)
    pdf.drawString(100, 800, "Damodar Grocery Store Invoice")
    pdf.drawString(100, 770, f"Order ID: {order_col(order, 'id', 0)}")
    pdf.drawString(100, 750, f"Customer: {order_col(order, 'customer_name', 1)}")
    pdf.drawString(100, 730, f"Phone: {order_col(order, 'phone', 2)}")
    pdf.drawString(100, 710, f"Address: {order_col(order, 'address', 3)}")

    y = 670

    for item in items:
        item_name = order_col(item, "product_name", 2)
        item_price = order_col(item, "price", 3)
        item_qty = order_col(item, "quantity", 4)
        pdf.drawString(100, y, f"{item_name} | Price: Rs.{item_price} | Qty: {item_qty}")
        y -= 20

    pdf.drawString(100, y - 20, f"Grand Total: Rs.{order_col(order, 'total', 4)}")
    pdf.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"invoice_{order_id}.pdf",
        mimetype="application/pdf"
    )


def format_order_datetime(value):
    raw = str(value or "").strip()
    missing = {
        "placed_at": "Order time not saved",
        "placed_date": "Date not saved",
        "placed_time": "Time not saved",
    }
    if not raw:
        return missing

    normalized = raw.replace("T", " ").replace("Z", "").split(".")[0]
    parsed = None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            parsed = datetime.strptime(normalized[:19], fmt)
            break
        except ValueError:
            continue

    if not parsed:
        return missing

    placed_ist = parsed + timedelta(hours=5, minutes=30)
    return {
        "placed_at": placed_ist.strftime("%d %b %Y, %I:%M %p"),
        "placed_date": placed_ist.strftime("%d %b %Y"),
        "placed_time": placed_ist.strftime("%I:%M %p"),
    }


def dashboard_order_stats(all_orders):
    total_orders = len(all_orders)
    total_sales = sum(int(order_col(order, "total", 4, 0) or 0) for order in all_orders)
    delivered_orders = sum(1 for order in all_orders if order_col(order, "status", 8) == "Delivered")
    cancelled_orders = sum(1 for order in all_orders if order_col(order, "status", 8) == "Cancelled")
    pending_orders = sum(
        1 for order in all_orders
        if order_col(order, "status", 8, "Pending") not in ["Delivered", "Cancelled"]
    )
    delivered_percent = round((delivered_orders / total_orders) * 100) if total_orders else 0
    cancelled_percent = round((cancelled_orders / total_orders) * 100) if total_orders else 0
    pending_percent = max(0, 100 - delivered_percent - cancelled_percent) if total_orders else 0

    return {
        "total_orders": total_orders,
        "total_sales": total_sales,
        "pending_orders": pending_orders,
        "delivered_orders": delivered_orders,
        "cancelled_orders": cancelled_orders,
        "delivered_percent": delivered_percent,
        "pending_percent": pending_percent,
        "cancelled_percent": cancelled_percent,
        "sales_percent": 100 if total_sales else 0,
        "orders_percent": 100 if total_orders else 0,
    }


@app.route("/orders")
def orders():
    if not is_admin():
        return redirect("/admin_login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM orders ORDER BY id DESC")
    all_orders = cursor.fetchall()
    stats = dashboard_order_stats(all_orders)

    order_items_map = {}
    order_meta_map = {}
    for order in all_orders:
        order_id = order_col(order, "id", 0)
        placed_meta = format_order_datetime(order_col(order, "created_at", 10, ""))
        order_meta_map[order_id] = placed_meta

        execute(cursor, "SELECT * FROM order_items WHERE order_id=?", (order_id,))
        items = []
        for item in cursor.fetchall():
            items.append({
                "product_id": order_col(item, "product_id", 5, ""),
                "product_name": order_col(item, "product_name", 2, "Product"),
                "price": order_col(item, "price", 3, 0),
                "quantity": order_col(item, "quantity", 4, 0),
            })
        order_items_map[order_id] = items

    conn.close()

    return render_template(
        "orders.html",
        orders=all_orders,
        order_items_map=order_items_map,
        order_meta_map=order_meta_map,
        **stats
    )


@app.route("/orders_stats")
def orders_stats():
    if not is_admin():
        return jsonify({"error": "Unauthorized"}), 401

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM orders")
    all_orders = cursor.fetchall()
    stats = dashboard_order_stats(all_orders)
    conn.close()

    return jsonify(stats)


@app.route("/update_status/<int:order_id>/<status>", methods=["GET", "POST"])
def update_status(order_id, status):
    if not is_admin():
        return redirect("/admin_login")

    if request.method == "POST":
        status = request.form.get("status", status)

    allowed_status = ["Pending", "Packed", "Shipped", "Delivered", "Cancelled"]
    if status not in allowed_status:
        return "Invalid status"

    conn = get_db()
    cursor = conn.cursor()
    execute(cursor, "UPDATE orders SET status=? WHERE id=?", (status, order_id))
    conn.commit()
    conn.close()

    return redirect("/orders")


@app.route("/cancel_order/<int:order_id>")
def cancel_order(order_id):
    if "user" not in session:
        return redirect("/customer_login")

    conn = get_db()
    cursor = conn.cursor()
    user_email_or_name = session["user"]

    execute(
        cursor,
        "SELECT * FROM orders WHERE id=? AND (customer_email=? OR customer_name=?)",
        (order_id, user_email_or_name, user_email_or_name)
    )
    order = cursor.fetchone()

    if not order:
        conn.close()
        return redirect("/my_orders")

    status = order_col(order, "status", 8, "Pending") or "Pending"
    if status not in ["Pending", "Packed"]:
        conn.close()
        return "Order sirf Pending ya Packed status me cancel ho sakta hai."

    execute(cursor, "SELECT * FROM order_items WHERE order_id=?", (order_id,))
    items = cursor.fetchall()

    for item in items:
        product_id = order_col(item, "product_id", 5, "")
        quantity = int(order_col(item, "quantity", 4, 0) or 0)

        if product_id and quantity > 0:
            execute(
                cursor,
                "UPDATE products SET stock = stock + ? WHERE id=?",
                (quantity, product_id)
            )

    execute(cursor, "UPDATE orders SET status=? WHERE id=?", ("Cancelled", order_id))
    conn.commit()
    conn.close()

    return redirect("/my_orders")


@app.route("/export_orders")
def export_orders():
    if not is_admin():
        return redirect("/admin_login")

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM orders ORDER BY id DESC")
    orders = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    ws.append(["ID", "Name", "Phone", "Address", "Total", "Payment Method", "Payment Status", "GST", "Status", "Email"])

    for order in orders:
        ws.append([
            order_col(order, "id", 0),
            order_col(order, "customer_name", 1),
            order_col(order, "phone", 2),
            order_col(order, "address", 3),
            order_col(order, "total", 4),
            order_col(order, "payment_method", 5),
            order_col(order, "payment_status", 6),
            order_col(order, "gst_amount", 7),
            order_col(order, "status", 8),
            order_col(order, "customer_email", 9),
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="orders_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/add_product", methods=["GET", "POST"])
def add_product():
    if not is_admin():
        return redirect("/admin_login")

    ensure_product_columns()

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        category = request.form.get("category", "").strip()
        price = request.form.get("price", "").strip()
        stock = request.form.get("stock", "").strip()
        description = request.form.get("description", "").strip()

        image_files = request.files.getlist("images")
        selected_images = [file for file in image_files if file and file.filename]

        try:
            price_value = int(price)
            stock_value = int(stock)
        except ValueError:
            return render_template("add_product.html", error="Price aur stock valid number hone chahiye.", form_data=request.form)

        if not name or not category:
            return render_template("add_product.html", error="Product name aur category required hai.", form_data=request.form)

        if price_value < 1:
            return render_template("add_product.html", error="Price 1 se kam nahi ho sakta.", form_data=request.form)

        if stock_value < 0:
            return render_template("add_product.html", error="Stock 0 se kam nahi ho sakta.", form_data=request.form)

        if not selected_images:
            return render_template("add_product.html", error="Kam se kam 1 product image select kare.", form_data=request.form)

        if len(selected_images) > 5:
            return render_template("add_product.html", error="Maximum 5 images hi upload kar sakte hain.", form_data=request.form)

        images = []

        for file in selected_images:
            if file and file.filename:
                secure_filename(file.filename)
                upload_result = cloudinary.uploader.upload(file)
                images.append(upload_result["secure_url"])

        conn = get_db()
        cursor = conn.cursor()

        insert_product_record(
            cursor,
            name,
            category,
            price_value,
            stock_value,
            description,
            images
        )

        conn.commit()
        conn.close()

        return render_template("add_product.html", success=f"{name} product add ho gaya.", form_data={})

    return render_template("add_product.html", form_data={})


@app.route("/download_product_template")
def download_product_template():
    if not is_admin():
        return redirect("/admin_login")

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["name", "category", "price", "stock", "description", "image", "image2", "image3", "image4", "image5"])
    ws.append(["HP Laptop Backpack", "Electronics", 999, 10, "Durable laptop backpack", "", "", "", "", ""])
    ws.append(["Face Cream", "Beauty", 199, 25, "Daily use skin cream", "", "", "", "", ""])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="product_upload_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/bulk_add_products", methods=["GET", "POST"])
def bulk_add_products():
    if not is_admin():
        return redirect("/admin_login")

    ensure_product_columns()

    if request.method == "POST":
        upload = request.files.get("product_file")

        if not upload or not upload.filename:
            return render_template("bulk_add_products.html", error="CSV ya Excel file select kare.")

        filename = upload.filename.lower()
        rows = []

        try:
            if filename.endswith(".xlsx"):
                workbook = load_workbook(upload.stream, data_only=True)
                sheet = workbook.active
                headers = [str(cell.value or "").strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    rows.append({
                        headers[index]: value
                        for index, value in enumerate(row)
                        if index < len(headers)
                    })
            elif filename.endswith(".csv"):
                text = upload.read().decode("utf-8-sig")
                reader = csv.DictReader(text.splitlines())
                rows = [row for row in reader if any((value or "").strip() for value in row.values())]
            else:
                return render_template("bulk_add_products.html", error="Sirf .xlsx ya .csv file upload kare.")
        except Exception:
            return render_template("bulk_add_products.html", error="File read nahi ho payi. Template format check kare.")

        if not rows:
            return render_template("bulk_add_products.html", error="File me product rows nahi mile.")

        conn = get_db()
        cursor = conn.cursor()
        added = 0
        skipped = 0
        errors = []

        for row_number, row in enumerate(rows, start=2):
            name = str(row.get("name") or "").strip()
            category = str(row.get("category") or "").strip()
            description = str(row.get("description") or "").strip()

            try:
                price = int(float(row.get("price") or 0))
                stock = int(float(row.get("stock") or 0))
            except (TypeError, ValueError):
                skipped += 1
                errors.append(f"Row {row_number}: price/stock invalid")
                continue

            if not name or not category or price < 1 or stock < 0:
                skipped += 1
                errors.append(f"Row {row_number}: required data missing")
                continue

            images = [
                str(row.get(field) or "").strip()
                for field in ["image", "image2", "image3", "image4", "image5"]
                if str(row.get(field) or "").strip()
            ]

            try:
                insert_product_record(cursor, name, category, price, stock, description, images)
                added += 1
            except Exception:
                skipped += 1
                errors.append(f"Row {row_number}: save failed")

        conn.commit()
        conn.close()

        return render_template(
            "bulk_add_products.html",
            success=f"{added} products add ho gaye.",
            skipped=skipped,
            errors=errors[:10]
        )

    return render_template("bulk_add_products.html")


@app.route("/edit_product/<int:product_id>", methods=["GET", "POST"])
def edit_product(product_id):
    if not is_admin():
        return redirect("/admin_login")

    ensure_product_columns()

    conn = get_db()
    cursor = conn.cursor()

    execute(cursor, "SELECT * FROM products WHERE id=?", (product_id,))
    product = cursor.fetchone()

    if not product:
        conn.close()
        return "Product not found"

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        category = request.form.get("category", "").strip()
        price = request.form.get("price", "").strip()
        stock = request.form.get("stock", "").strip()
        description = request.form.get("description", "").strip()

        try:
            price_value = int(price)
            stock_value = int(stock)
        except ValueError:
            conn.close()
            return render_template("edit_product.html", product=product, error="Price aur stock valid number hone chahiye.")

        if not name or not category:
            conn.close()
            return render_template("edit_product.html", product=product, error="Product name aur category required hai.")

        if price_value < 1:
            conn.close()
            return render_template("edit_product.html", product=product, error="Price 1 se kam nahi ho sakta.")

        if stock_value < 0:
            conn.close()
            return render_template("edit_product.html", product=product, error="Stock 0 se kam nahi ho sakta.")

        old_images = [
            product_col(product, "image", 4, ""),
            product_col(product, "image2", 7, ""),
            product_col(product, "image3", 8, ""),
            product_col(product, "image4", 9, ""),
            product_col(product, "image5", 10, ""),
        ]

        fields = ["image1", "image2", "image3", "image4", "image5"]
        new_images = []

        for index, field_name in enumerate(fields):
            file = request.files.get(field_name)
            remove_image = request.form.get(f"remove_{field_name}") == "1"

            if file and file.filename:
                secure_filename(file.filename)
                upload_result = cloudinary.uploader.upload(file)
                new_images.append(upload_result["secure_url"])
            elif remove_image:
                new_images.append("")
            else:
                new_images.append(old_images[index])

        execute(
            cursor,
            """
            UPDATE products
            SET name=?,
                category=?,
                price=?,
                image=?,
                stock=?,
                description=?,
                image2=?,
                image3=?,
                image4=?,
                image5=?
            WHERE id=?
            """,
            (
                name,
                category,
                price_value,
                new_images[0],
                stock_value,
                description,
                new_images[1],
                new_images[2],
                new_images[3],
                new_images[4],
                product_id
            )
        )

        conn.commit()

        execute(cursor, "SELECT * FROM products WHERE id=?", (product_id,))
        updated_product = cursor.fetchone()

        conn.close()

        return render_template("edit_product.html", product=updated_product, success="Product update ho gaya.")

    conn.close()
    return render_template("edit_product.html", product=product)


@app.route("/delete_product/<int:product_id>")
def delete_product(product_id):
    if not is_admin():
        return redirect("/admin_login")

    conn = get_db()
    cursor = conn.cursor()
    execute(cursor, "DELETE FROM products WHERE id=?", (product_id,))
    conn.commit()
    conn.close()

    return redirect("/")


@app.route("/bulk_delete_products", methods=["POST"])
def bulk_delete_products():
    if not is_admin():
        return redirect("/admin_login")

    product_ids = []
    for value in request.form.getlist("product_ids"):
        try:
            product_ids.append(int(value))
        except (TypeError, ValueError):
            pass

    if not product_ids:
        return redirect("/")

    placeholders = ",".join(["?"] * len(product_ids))
    conn = get_db()
    cursor = conn.cursor()

    execute(cursor, f"DELETE FROM wishlist WHERE product_id IN ({placeholders})", product_ids)
    execute(cursor, f"DELETE FROM reviews WHERE product_id IN ({placeholders})", product_ids)
    execute(cursor, f"DELETE FROM products WHERE id IN ({placeholders})", product_ids)

    conn.commit()
    conn.close()

    return redirect("/")


if __name__ == "__main__":
    app.run(debug=True)
